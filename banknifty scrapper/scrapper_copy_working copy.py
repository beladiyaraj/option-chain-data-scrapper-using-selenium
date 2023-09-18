import json
import time
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import NamedStyle


with open('config.json', 'r') as f:
    config = json.load(f)

chrome_options = webdriver.ChromeOptions()
chrome_options.binary_location = "GoogleChromePortable64/App/Chrome-bin/chrome.exe"
chrome_options.add_argument("--log-level=3")

service = Service(config['chrome_driver_path'])
driver = webdriver.Chrome(service=service, options=chrome_options)

login_url = "https://web.quantsapp.com/signin"
driver.get(login_url)

wait = WebDriverWait(driver, 300)

wait.until(EC.url_contains("https://web.quantsapp.com/home"))

time.sleep(5)

# Main URL
base_url = config['base_url']
option_type = config['optionType']
expiry_dates = config['expiryDates']

# Create a new Excel workbook
workbook = Workbook()
default_sheet = workbook.active
workbook.remove(default_sheet)
output_filename = config['output_filename']

header_row = ["GAMMA", "THEETA", "DELTA", "VEGA", "IV", "OI", "COI", "VOL", "LTP", "CLTP",
              "Strike", "CLTP", "LTP", "VOL", "COI", "OI", "IV", "VEGA", "DELTA", "THEETA", "Gamma"]

number_style = NamedStyle(name='number_style')
number_style.number_format = '0.00'

while True:
    # Loop through the expiry dates
    max_retries = 3
    for expiry in expiry_dates:
        url = f"{base_url}{option_type}?symbol={option_type}&expiry={expiry}"
        print(f"Accessing URL: {url}")
        driver.get(url)

        # Adjust this XPath to target the <tr> elements within the <tbody>
        xpath_query_rows = "//div[@id='dataTable']/table/tbody/tr"

        # Loop until no new tr elements are loading or a maximum timeout is reached
        previous_count = 0
        retry_count = 0  # Initialize the retry counter
        while retry_count < max_retries:
            try:
                wait.until(EC.presence_of_element_located(
                    (By.XPATH, xpath_query_rows)))

                rows = driver.find_elements(By.XPATH, xpath_query_rows)
                if len(rows) == previous_count:
                    break  # Break out of the loop if elements have loaded successfully

                previous_count = len(rows)
            except Exception as e:
                retry_count += 1
                print(
                    f"Retry attempt {retry_count} for expiry {expiry} failed. Error: {e}")

            if retry_count == max_retries:
                print(
                    f"Max retry attempts reached for expiry {expiry}. Moving on to the next expiry.")
                continue  # Move on to the next expiry date if max retries are reachedk

        # Extract HTML content from rows
        html_data_list = []

        for row in rows:
            try:
                # Extract the HTML content from the row
                html_content = row.get_attribute("innerHTML")

                # Parse the HTML content with BeautifulSoup
                soup = BeautifulSoup(html_content, 'html.parser')

                # Find all <span> elements with the text "CHECK" and remove them from the soup
                check_spans = soup.find_all('span', string='CHECK')
                for check_span in check_spans:
                    check_span.decompose()

                # Add the modified HTML content (without the "CHECK" spans) to the html_data_list
                html_data_list.append([str(soup)])
            except Exception as e:
                print(f"Failed to extract HTML content from a row due to {e}")

        # Extract text from each "bunch" of <td> elements and save in the same row
        text_data_list = []

        for i in range(0, len(html_data_list), 2):
            try:
                # Parse the HTML content with BeautifulSoup
                soup1 = BeautifulSoup(html_data_list[i][0], 'html.parser')
                soup2 = BeautifulSoup(html_data_list[i+1][0], 'html.parser')

                # Extract text from each <td> element in the "bunch" and remove empty strings
                text_list1 = [td.get_text(strip=True) for td in soup1.find_all(
                    'td') if td.get_text(strip=True)]
                text_list2 = [td.get_text(strip=True) for td in soup2.find_all(
                    'td') if td.get_text(strip=True)]

                # Create a copy of the original list
                modified_list_of_text_list_1 = []

                # Iterate through the original text and split values at positions 0, 3, 5, and 8
                positions_to_split = [0, 3, 5, 8]

                for i, text in enumerate(text_list1):
                    if i in positions_to_split:
                        if '(' in text and ')' in text:
                            parts = text.split('(')
                            something1 = parts[0]
                            something2 = parts[1].strip(')')

                            # Append the split values
                            modified_list_of_text_list_1.append(something1)
                            modified_list_of_text_list_1.append(something2)
                        else:
                            # If no parentheses are found, append the original value and an empty string
                            modified_list_of_text_list_1.append(text)
                            modified_list_of_text_list_1.append('0')
                    else:
                        modified_list_of_text_list_1.append(text)

                positions_to_remove_hyphen = [4, 7]

                for i in positions_to_remove_hyphen:
                    if i < len(modified_list_of_text_list_1) and '-' in modified_list_of_text_list_1[i]:
                        modified_list_of_text_list_1[i] = modified_list_of_text_list_1[i].replace(
                            '-', '')

                numeric_pattern = r"[-+]?\d*\.\d+|\d+"

                # Extract numeric values using the regex pattern
                modified_list_of_text_list_2 = []

                for text in text_list2:
                    text = text.replace("-", "0")
                    match = re.findall(numeric_pattern, text)
                    if match:
                        value = match[0]
                        if "." in value:
                            modified_list_of_text_list_2.append(float(value))
                        else:
                            modified_list_of_text_list_2.append(int(value))

                # Print the resulting modified list
                merged_text_list = modified_list_of_text_list_1 + modified_list_of_text_list_2
                text_data_list.append(merged_text_list)

            except Exception as e:
                print(f"Failed to extract text from HTML content due to {e}")

        gamma1_column = []
        theta1_column = []
        delta1_column = []
        vega1_column = []
        iv1_column = []
        oi1_column = []
        coi1_column = []
        vol1_column = []
        ltp1_column = []
        strike_column = []
        gamma2_column = []
        theta2_column = []
        delta2_column = []
        vega2_column = []
        iv2_column = []
        coi2_column = []
        oi2_column = []
        vol2_column = []
        ltp2_column = []
        cltp1_column = []
        cltp2_column = []

        # Loop through the extracted data and populate the corresponding lists
        for item in text_data_list:
            gamma1_column.append(item[14])
            theta1_column.append(item[15])
            delta1_column.append(item[16])
            vega1_column.append(item[13])
            iv1_column.append(item[3])
            oi1_column.append(item[0])
            coi1_column.append(item[1])
            vol1_column.append(item[2])
            ltp1_column.append(item[4])
            cltp1_column.append(item[5])
            strike_column.append(item[6])
            cltp2_column.append(item[8])
            ltp2_column.append(item[7])
            vol2_column.append(item[10])
            coi2_column.append(item[12])
            oi2_column.append(item[11])
            iv2_column.append(item[9])
            vega2_column.append(item[20])
            delta2_column.append(item[17])
            theta2_column.append(item[18])
            gamma2_column.append(item[19])

        def convert_str_to_number_list(str_list):
            numeric_list = []
            for item in str_list:
                if isinstance(item, str):
                    item = item.replace(',', '')  # Remove commas from strings
                    try:
                        numeric_value = float(item)  # Try to convert to float
                    except ValueError:
                        numeric_value = item  # If conversion fails, keep it as a string
                    numeric_list.append(numeric_value)
                else:
                    numeric_list.append(item)  # Keep non-string elements as is
            return numeric_list

        # Apply the function to all the lists
        gamma1_column = convert_str_to_number_list(gamma1_column)
        theta1_column = convert_str_to_number_list(theta1_column)
        delta1_column = convert_str_to_number_list(delta1_column)
        vega1_column = convert_str_to_number_list(vega1_column)
        iv1_column = convert_str_to_number_list(iv1_column)
        oi1_column = convert_str_to_number_list(oi1_column)
        coi1_column = convert_str_to_number_list(coi1_column)
        vol1_column = convert_str_to_number_list(vol1_column)
        ltp1_column = convert_str_to_number_list(ltp1_column)
        strike_column = convert_str_to_number_list(strike_column)
        gamma2_column = convert_str_to_number_list(gamma2_column)
        theta2_column = convert_str_to_number_list(theta2_column)
        delta2_column = convert_str_to_number_list(delta2_column)
        vega2_column = convert_str_to_number_list(vega2_column)
        iv2_column = convert_str_to_number_list(iv2_column)
        coi2_column = convert_str_to_number_list(coi2_column)
        oi2_column = convert_str_to_number_list(oi2_column)
        vol2_column = convert_str_to_number_list(vol2_column)
        ltp2_column = convert_str_to_number_list(ltp2_column)
        cltp1_column = convert_str_to_number_list(cltp1_column)
        cltp2_column = convert_str_to_number_list(cltp2_column)

        # Create a new sheet with the expiry date as the name
        if expiry in workbook.sheetnames:
            sheet = workbook[expiry]  # Get the existing sheet
        else:
            sheet = workbook.create_sheet(
                expiry, index=len(workbook.sheetnames))

        # Convert the text_data_list to a DataFrame
        for idx, col_name in enumerate(header_row, 1):
            sheet.cell(row=1, column=idx, value=col_name)

        # Write the DataFrame to the sheet in columns
        for row_idx in range(2, len(gamma1_column) + 2):
            sheet.cell(row=row_idx, column=1, value=gamma1_column[row_idx - 2])
            sheet.cell(row=row_idx, column=2, value=theta1_column[row_idx - 2])
            sheet.cell(row=row_idx, column=3, value=delta1_column[row_idx - 2])
            sheet.cell(row=row_idx, column=4, value=vega1_column[row_idx - 2])
            sheet.cell(row=row_idx, column=5, value=iv1_column[row_idx - 2])
            sheet.cell(row=row_idx, column=6, value=oi1_column[row_idx - 2])
            sheet.cell(row=row_idx, column=7, value=coi1_column[row_idx - 2])
            sheet.cell(row=row_idx, column=8, value=vol1_column[row_idx - 2])
            sheet.cell(row=row_idx, column=9, value=ltp1_column[row_idx - 2])
            sheet.cell(row=row_idx, column=10, value=cltp1_column[row_idx - 2])
            sheet.cell(row=row_idx, column=11,
                       value=strike_column[row_idx - 2])
            sheet.cell(row=row_idx, column=12, value=cltp2_column[row_idx - 2])
            sheet.cell(row=row_idx, column=13, value=ltp2_column[row_idx - 2])
            sheet.cell(row=row_idx, column=14, value=vol2_column[row_idx - 2])
            sheet.cell(row=row_idx, column=15, value=coi2_column[row_idx - 2])
            sheet.cell(row=row_idx, column=16, value=oi2_column[row_idx - 2])
            sheet.cell(row=row_idx, column=17, value=iv2_column[row_idx - 2])
            sheet.cell(row=row_idx, column=18, value=vega2_column[row_idx - 2])
            sheet.cell(row=row_idx, column=19,
                       value=delta2_column[row_idx - 2])
            sheet.cell(row=row_idx, column=20,
                       value=theta2_column[row_idx - 2])
            sheet.cell(row=row_idx, column=21,
                       value=gamma2_column[row_idx - 2])

        try:
            workbook.save(output_filename)
        except Exception as e:
            print(f"Exception occurred while saving the workbook: {e}")
            time.sleep(1)  # Sleep for 1 second before retrying
        else:
            # If save is successful, break out of the loop
            break